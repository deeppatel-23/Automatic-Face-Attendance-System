from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_session import Session
from datetime import datetime, date
import os
import glob
from werkzeug.utils import secure_filename
from config.config import Config
from database.db_manager import db
from face_recognition.face_detector import FaceRecognizer
from email_service.email_sender import email_sender
import cv2
import base64
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import pandas as pd

app = Flask(__name__)
app.config.from_object(Config)

# Initialize session
Session(app)

# Initialize face recognizer
face_recognizer = FaceRecognizer(Config.FACE_ENCODINGS_PATH, Config.TOLERANCE, Config.MODEL)

# Utility function to check allowed file
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS

def get_current_session():
    """Get current attendance session"""
    current_time = datetime.now().time()
    sessions = Config.ATTENDANCE_SESSIONS
    
    for session_name, times in sessions.items():
        start = datetime.strptime(times['start'], '%H:%M').time()
        end = datetime.strptime(times['end'], '%H:%M').time()
        if start <= current_time <= end:
            return session_name
    
    return None

# Routes

@app.route('/')
def home():
    """Home page"""
    if 'user_id' in session:
        user = db.get_user_by_id(session['user_id'])
        if user['role'] == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if request.method == 'POST':
        # Get input - can be either email or student_id
        identifier = request.form.get('email')  # Field is still named 'email' but accepts both
        password = request.form.get('password')
        
        # Try to get user by email OR student_id
        user = db.get_user_by_email_or_student_id(identifier)
        
        if user and db.verify_password(user['password'], password):
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            session['role'] = user['role']
            
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('student_dashboard'))
        else:
            flash('Invalid email, student ID, or password', 'danger')
    
    return render_template('login_page.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Registration page"""
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        # Validation
        if not all([student_id, name, email, password, confirm_password]):
            if is_ajax:
                return jsonify({'success': False, 'message': 'All fields are required'}), 400
            flash('All fields are required', 'danger')
            return render_template('register_page.html')

        if password != confirm_password:
            if is_ajax:
                return jsonify({'success': False, 'message': 'Passwords do not match'}), 400
            flash('Passwords do not match', 'danger')
            return render_template('register_page.html')

        # Check uploaded photos (expect multiple)
        photos = request.files.getlist('photos')
        valid_photos = [f for f in photos if f and f.filename and allowed_file(f.filename)]
        if len(valid_photos) < 5:
            if is_ajax:
                return jsonify({'success': False, 'message': 'Please upload at least 5 clear face photos.'}), 400
            flash('Please upload at least 5 clear face photos (different angles/expressions).', 'danger')
            return render_template('register_page.html')

        # Register user
        user_id = db.register_user(student_id, name, email, password)

        if user_id:
            # Ensure upload folder exists
            os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)

            # Save all photos and register encodings
            for idx, file in enumerate(valid_photos, start=1):
                try:
                    filename = secure_filename(f"{user_id}_{idx}_{file.filename}")
                    filepath = os.path.join(Config.UPLOAD_FOLDER, filename)
                    file.save(filepath)

                    # Extract and register face encoding for each sample
                    encoding = face_recognizer.register_face(user_id, filepath)
                    if encoding:
                        # Store latest encoding path in DB (keeps photo_path and face_encoding)
                        db.update_user_face_encoding(user_id, str(encoding), filepath)
                except Exception:
                    # ignore single-file failures, continue with others
                    continue

            # Send welcome email
            email_sender.send_welcome_email(email, name)

            if is_ajax:
                return jsonify({
                    'success': True,
                    'message': 'Registration successful! Please login.',
                    'redirect': url_for('login')
                }), 200
            flash('Registration successful! Please login.', 'success')
            return redirect(url_for('login'))
        else:
            if is_ajax:
                return jsonify({'success': False, 'message': 'Email or Student ID already registered'}), 400
            flash('Email or Student ID already registered', 'danger')
    
    return render_template('register_page.html')

@app.route('/logout')
def logout():
    """Logout"""
    session.clear()
    flash('Logged out successfully', 'success')
    return redirect(url_for('home'))

@app.route('/student/dashboard')
def student_dashboard():
    """Student dashboard"""
    if 'user_id' not in session or session.get('role') != 'student':
        return redirect(url_for('login'))
    
    user = db.get_user_by_id(session['user_id'])
    stats = db.get_attendance_statistics(session['user_id'])
    current_session = get_current_session()
    
    # Get recent attendance
    recent_attendance = db.get_user_attendance(session['user_id'])[:10]
    
    return render_template('student_dashboard.html',
                         user=user,
                         stats=stats,
                         current_session=current_session,
                         recent_attendance=recent_attendance)

@app.route('/admin/dashboard')
def admin_dashboard():
    """Admin dashboard"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    total_students = len(db.get_all_students())
    today_attendance = db.get_attendance_by_date(str(date.today()))
    current_session = get_current_session()
    pending_requests = db.get_pending_attendance_requests(str(date.today()))
    
    return render_template('admin_dashboard.html',
                         total_students=total_students,
                         today_attendance=len(today_attendance),
                         current_session=current_session,
                         pending_requests=len(pending_requests))

@app.route('/admin/pending_attendance')
def admin_pending_attendance():
    """Admin can approve/reject pending face scan attendance requests"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))

    pending_requests = db.get_pending_attendance_requests(str(date.today()))
    return render_template('admin_pending_attendance.html', pending_requests=pending_requests)

@app.route('/admin/approve_attendance/<int:request_id>', methods=['POST'])
def admin_approve_attendance(request_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    request_record = db.get_attendance_request(request_id)
    if not request_record:
        return jsonify({'success': False, 'message': 'Request not found'}), 404

    # Mark attendance and update request status
    if db.mark_attendance(request_record['user_id'], request_record['date'], request_record['session'], 'present'):
        db.update_attendance_request_status(request_id, 'approved', processed_by=session['user_id'])
        return jsonify({'success': True, 'message': 'Attendance approved and marked present'}), 200
    return jsonify({'success': False, 'message': 'Failed to mark attendance'}), 500

@app.route('/admin/reject_attendance/<int:request_id>', methods=['POST'])
def admin_reject_attendance(request_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    request_record = db.get_attendance_request(request_id)
    if not request_record:
        return jsonify({'success': False, 'message': 'Request not found'}), 404

    if db.update_attendance_request_status(request_id, 'rejected', processed_by=session['user_id']):
        return jsonify({'success': True, 'message': 'Attendance request rejected'}), 200
    return jsonify({'success': False, 'message': 'Failed to reject attendance request'}), 500

@app.route('/admin/students')
def admin_students():
    """View all students"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    students = db.get_all_students()
    app_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    for student in students:
        photo_path = student.get('photo_path')
        if photo_path:
            if os.path.isabs(photo_path):
                try:
                    rel = os.path.relpath(photo_path, app_root).replace('\\', '/')
                    normalized = '/' + rel
                except Exception:
                    normalized = '/' + photo_path.replace('\\', '/').lstrip('/')
            else:
                normalized = '/' + photo_path.replace('\\', '/').lstrip('/')

            student['photo_url'] = normalized

            # If DB has an absolute or raw path, normalize and store it
            if not photo_path.startswith('/') or photo_path.startswith(app_root.replace('\\', '/')):
                new_db_path = normalized.lstrip('/')
                if new_db_path != photo_path:
                    db.update_user_photo_path(student['id'], new_db_path)
                    student['photo_path'] = new_db_path
        else:
            student['photo_url'] = None

    return render_template('admin_students.html', students=students)

@app.route('/admin/student/<int:student_id>')
def admin_student_profile(student_id):
    """View individual student profile"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    student = db.get_user_by_id(student_id)
    if not student or student['role'] != 'student':
        flash('Student not found', 'danger')
        return redirect(url_for('admin_students'))
    
    # Get attendance records
    attendance = db.get_user_attendance(student_id)
    stats = db.get_attendance_statistics(student_id)
    current_session = get_current_session()

    # Determine session and status for toggle button
    today = str(date.today())
    session_key = current_session or 'manual'
    record = db.get_attendance_record(student_id, today, session_key)
    session_status = record['status'] if record else 'absent'
    display_session = current_session if current_session else 'manual'

    # Get all photos for this student and convert to URL-safe relative paths
    student_photos_abs = glob.glob(os.path.join(Config.UPLOAD_FOLDER, f"{student_id}_*.jpg")) + \
                    glob.glob(os.path.join(Config.UPLOAD_FOLDER, f"{student_id}_*.png"))

    student_photos = []
    app_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    for p in student_photos_abs:
        try:
            rel = os.path.relpath(p, app_root).replace('\\', '/')
            student_photos.append('/' + rel)
        except Exception:
            student_photos.append('/' + p.replace('\\', '/').lstrip('/'))

    # Ensure student.photo_path is a valid URL for the main view
    photo_url = None
    photo_path = student.get('photo_path')
    if photo_path:
        if os.path.isabs(photo_path):
            rel = os.path.relpath(photo_path, app_root).replace('\\', '/')
            photo_url = '/' + rel
        else:
            photo_url = '/' + photo_path.replace('\\', '/').lstrip('/')
    student['photo_url'] = photo_url

    return render_template('admin_student_profile.html',
                         student=student,
                         attendance=attendance,
                         stats=stats,
                         current_session=current_session,
                         session_status=session_status,
                         display_session=display_session,
                         photos=student_photos,
                         date=date,
                         config=Config)


@app.route('/admin/toggle_student_status/<int:student_id>', methods=['POST'])
def admin_toggle_student_status(student_id):
    """Admin toggles presence/absence for a student in current session"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    session_key = get_current_session() or 'manual'

    student = db.get_user_by_id(student_id)
    if not student or student.get('role') != 'student':
        return jsonify({'success': False, 'message': 'Student not found'}), 404

    try:
        today = str(date.today())
        record = db.get_attendance_record(student_id, today, session_key)

        if record and record.get('status') == 'present':
            new_status = 'absent'
        else:
            new_status = 'present'

        db.mark_attendance(student_id, today, session_key, new_status)
        return jsonify({'success': True, 'message': f"{student.get('name')} marked {new_status} for {session_key}", 'status': new_status}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/manual_attendance/<int:student_id>', methods=['POST'])
def admin_manual_attendance(student_id):
    """Admin sets attendance status for arbitrary date/session"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    try:
        data = request.get_json() or {}
        attendance_date = data.get('date') or str(date.today())
        session_name = data.get('session')
        status = data.get('status', 'present').lower()

        if status not in ['present', 'absent']:
            return jsonify({'success': False, 'message': 'Invalid status. Use present or absent.'}), 400

        if not session_name or session_name not in Config.ATTENDANCE_SESSIONS.keys():
            return jsonify({'success': False, 'message': 'Invalid session name.'}), 400

        student = db.get_user_by_id(student_id)
        if not student or student.get('role') != 'student':
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        db.mark_attendance(student_id, attendance_date, session_name, status)
        return jsonify({'success': True, 'message': f"{student.get('name')} marked {status} for {session_name} on {attendance_date}", 'status': status}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/mark_present/<int:student_id>', methods=['POST'])
def admin_mark_present(student_id):
    """Admin marks individual student present for the current session"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    session_key = get_current_session() or 'manual'

    try:
        student = db.get_user_by_id(student_id)
        if not student or student.get('role') != 'student':
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        today = str(date.today())
        db.mark_attendance(student_id, today, session_key, 'present')
        return jsonify({'success': True, 'message': f'{student.get("name")} marked present for {session_key}'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/remove_student/<int:student_id>', methods=['DELETE', 'POST'])
def remove_student(student_id):
    """Remove a student from the system"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        # Prevent admin from deleting themselves
        if student_id == session['user_id']:
            return jsonify({'success': False, 'message': 'Cannot delete your own account'}), 400
        
        # Delete student using database method
        if db.delete_student(student_id):
            return jsonify({'success': True, 'message': 'Student removed successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to remove student'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/bulk_import', methods=['GET', 'POST'])
def bulk_import():
    """Bulk import students from CSV"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        if 'csv_file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(url_for('bulk_import'))
        
        file = request.files['csv_file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('bulk_import'))
        
        if not file.filename.endswith('.csv'):
            flash('Only CSV files allowed', 'danger')
            return redirect(url_for('bulk_import'))
        
        try:
            # Read CSV
            df = pd.read_csv(file)
            
            # Validate columns
            required_cols = ['student_id', 'name', 'email', 'password']
            if not all(col in df.columns for col in required_cols):
                flash(f'CSV must contain columns: {", ".join(required_cols)}', 'danger')
                return redirect(url_for('bulk_import'))
            
            # Import students
            success_count = 0
            error_count = 0
            
            for idx, row in df.iterrows():
                try:
                    user_id = db.register_user(
                        row['student_id'].strip(),
                        row['name'].strip(),
                        row['email'].strip(),
                        row['password'].strip()
                    )
                    
                    if user_id:
                        # Send welcome email
                        email_sender.send_welcome_email(row['email'].strip(), row['name'].strip())
                        success_count += 1
                    else:
                        error_count += 1
                except Exception as e:
                    error_count += 1
            
            flash(f'✅ {success_count} students imported successfully! ❌ {error_count} failed.', 'success')
            return redirect(url_for('admin_students'))
        
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'danger')
            return redirect(url_for('bulk_import'))
    
    return render_template('bulk_import.html')

@app.route('/admin/download_csv_template')
def download_csv_template():
    """Download CSV template for bulk import"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    
    # Create sample CSV
    df = pd.DataFrame({
        'student_id': ['STU001', 'STU002', 'STU003'],
        'name': ['John Doe', 'Jane Smith', 'Bob Johnson'],
        'email': ['john@example.com', 'jane@example.com', 'bob@example.com'],
        'password': ['pass123', 'pass456', 'pass789']
    })
    
    output = BytesIO()
    df.to_csv(output, index=False)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='text/csv',
        as_attachment=True,
        download_name='student_import_template.csv'
    )

@app.route('/admin/mark_absent', methods=['POST'])
def admin_mark_absent():
    """Mark absent students and send emails"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        session_name = request.json.get('session')
        if not session_name:
            return jsonify({'success': False, 'message': 'Session not specified'})
        
        today = str(date.today())
        absent_students = db.get_absent_students(today, session_name)
        
        # Mark as absent and collect emails
        email_list = []
        for student in absent_students:
            db.mark_attendance(student['id'], today, session_name, 'absent')
            email_list.append({
                'email': student['email'],
                'name': student['name'],
                'session': session_name
            })
        
        # Send bulk emails
        results = email_sender.send_bulk_absence_notification(email_list)
        
        return jsonify({
            'success': True,
            'message': f'Marked {len(absent_students)} students absent',
            'emails_sent': results['success'],
            'emails_failed': results['failed']
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    """Mark attendance via face recognition"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'}), 401
    
    try:
        # Get image data from request
        image_data = request.json.get('image')
        if not image_data:
            return jsonify({'success': False, 'message': 'No image provided'})
        
        # Decode base64 image
        import base64
        image_bytes = base64.b64decode(image_data.split(',')[1])
        nparr = np.frombuffer(image_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        # Recognize face
        results = face_recognizer.recognize_face(frame)
        
        if results and results[0]['name'] != 'Unknown':
            user_id = int(results[0]['name'])
            if user_id == session['user_id']:
                # Add pending attendance request for admin approval
                today = str(date.today())
                current_session = get_current_session()

                if current_session:
                    if db.add_attendance_request(user_id, today, current_session):
                        user = db.get_user_by_id(user_id)
                        email_sender.send_attendance_confirmation(user['email'], user['name'], current_session)
                        return jsonify({
                            'success': True,
                            'message': 'Face recognized. Attendance request submitted for admin approval.'
                        })
                    else:
                        return jsonify({
                            'success': False,
                            'message': 'Failed to submit attendance request.'
                        })
                else:
                    return jsonify({
                        'success': False,
                        'message': 'No active session at this time'
                    })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Face does not match your account'
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Face not recognized. Please try again.'
            })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/attendance_report')
def attendance_report():
    """Attendance report page"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    return render_template('attendance_report.html')

@app.route('/api/attendance_data', methods=['GET'])
def get_attendance_data():
    """Get attendance data for reports"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': 'Date range required'})
        
        students = db.get_all_students()
        attendance_data = []
        
        for student in students:
            records = db.get_user_attendance(student['id'], start_date, end_date)
            attendance_data.append({
                'student_id': student['student_id'],
                'name': student['name'],
                'email': student['email'],
                'total_sessions': len(records),
                'present': sum(1 for r in records if r['status'] == 'present'),
                'absent': sum(1 for r in records if r['status'] == 'absent')
            })
        
        return jsonify({'success': True, 'data': attendance_data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/export_excel', methods=['GET'])
def export_excel():
    """Export attendance report to Excel"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': 'Date range required'})
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance'
        
        # Headers
        headers = ['Student ID', 'Name', 'Email', 'Total Sessions', 'Present', 'Absent', 'Percentage']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Get data
        students = db.get_all_students()
        for student in students:
            records = db.get_user_attendance(student['id'], start_date, end_date)
            total = len(records)
            present = sum(1 for r in records if r['status'] == 'present')
            absent = sum(1 for r in records if r['status'] == 'absent')
            percentage = (present / total * 100) if total > 0 else 0
            
            ws.append([
                student['student_id'],
                student['name'],
                student['email'],
                total,
                present,
                absent,
                f"{percentage:.2f}%"
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"attendance_{start_date}_to_{end_date}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    """Forgot password request"""
    if request.method == 'POST':
        email = request.form.get('email')
        user = db.get_user(email)
        
        if user:
            # Generate reset token
            import secrets
            from datetime import timedelta
            token = secrets.token_urlsafe(32)
            expiry = datetime.now() + timedelta(hours=1)
            db.set_reset_token(email, token, expiry)
            
            # Send reset email
            reset_link = f"{request.host_url.rstrip('/')}/reset_password/{token}"
            email_sender.send_password_reset_email(email, user['name'], reset_link)
            flash('Password reset link sent to your email (valid for 1 hour).', 'info')
        else:
            flash('If email exists, reset link will be sent.', 'info')
        
        return redirect(url_for('login'))
    
    return render_template('forgot_password.html')

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    """Reset password with token"""
    user = db.verify_reset_token(token)
    
    if not user:
        flash('Invalid or expired reset link.', 'danger')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if not password or password != confirm_password:
            flash('Passwords do not match or empty.', 'danger')
            return render_template('reset_password.html', token=token)
        
        if db.update_password(user['id'], password):
            db.clear_reset_token(user['id'])
            flash('Password reset successful! Please login with your new password.', 'success')
            return redirect(url_for('login'))
        else:
            flash('Error resetting password.', 'danger')
    
    return render_template('reset_password.html', token=token)

@app.route('/admin/reset_student_password', methods=['GET', 'POST'])
def admin_reset_password():
    """Admin reset student password"""
    if 'user_id' not in session or not is_admin(session['user_id']):
        flash('Unauthorized access', 'danger')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        new_password = request.form.get('new_password')
        
        if not student_id or not new_password:
            flash('Student ID and password required', 'danger')
            return redirect(url_for('admin_reset_password'))
        
        # Find user by student_id
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM users WHERE student_id = ?', (student_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            user_id = result[0]
            if db.update_password(user_id, new_password):
                flash(f'Password reset for {student_id} successfully!', 'success')
            else:
                flash('Error resetting password', 'danger')
        else:
            flash('Student not found', 'danger')
        
        return redirect(url_for('admin_reset_password'))
    
    # Get all students for dropdown
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, student_id, name FROM users WHERE role = ?', ('student',))
    students = cursor.fetchall()
    conn.close()
    
    return render_template('admin_reset_password.html', students=students)

def is_admin(user_id):
    """Check if user is admin"""
    user = db.get_user_by_id(user_id)
    return user and user['role'] == 'admin'

@app.errorhandler(404)
def page_not_found(e):
    """404 error handler"""
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    """500 error handler"""
    return render_template('500.html'), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
